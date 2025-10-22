"""
logger_excel.py
---------------
Excel logging and thumbnail management for MDA camera automation.

Includes:
- Automatic Excel creation with project metadata
- Thumbnails for each logged image
- Hyperlinks to files on disk
- Reordering by FEATURE_ORDER

Author: Syrine Bouhoula / MDA Project
Date: 2025
"""

import os
import time
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
from error_handler import ErrorHandler


class ExcelLogger:
    """Handles Excel creation, metadata, and image logging."""

    HEADERS = ["Order", "timestamp", "Feature", "Image", "tv", "av", "zoom", "iso", "ec"]

    def __init__(self, log_path, capture_dir, header_cfg=None, thumbnail_width=180, feature_order=None):
        self.log_path = log_path
        self.capture_dir = capture_dir
        self.thumb_width = thumbnail_width
        self.header_cfg = header_cfg or {}
        self.feature_order = feature_order or []
        self._init_excel()

    # ====== Internal setup ======
    def _init_excel(self):
        """Create or open Excel file with metadata + headers."""
        if os.path.exists(self.log_path):
            try:
                wb = load_workbook(self.log_path)
                ws = wb.active
                print(f"✅ Using existing Excel log: {self.log_path}")
                self.wb, self.ws = wb, ws
                return
            except Exception:
                pass

        wb = Workbook()
        ws = wb.active
        self._write_metadata(ws)
        ws.append(self.HEADERS)
        wb.save(self.log_path)
        self.wb, self.ws = wb, ws
        print(f"🆕 Created new Excel log at {self.log_path}")

    # ====== Metadata ======
    def _write_metadata(self, ws):
        """Write project metadata rows at the top of the Excel sheet."""
        ws["A1"] = "Part number"
        ws["B1"] = self.header_cfg.get("Part number", "")
        ws["A2"] = "Part Description"
        ws["B2"] = self.header_cfg.get("Part Description", "")
        ws["A3"] = "Serial number"
        ws["B3"] = self.header_cfg.get("Serial number", "")
        ws["A4"] = "Program type"
        ws["B4"] = self.header_cfg.get("Program type", "")

    # ====== Thumbnail generation ======
    def _make_thumbnail(self, src_path):
        """Generate thumbnail and save it as PNG."""
        if not os.path.exists(src_path):
            return None
        thumb_dir = os.path.join(self.capture_dir, "_thumbs")
        os.makedirs(thumb_dir, exist_ok=True)
        base = os.path.splitext(os.path.basename(src_path))[0]
        out_path = os.path.join(thumb_dir, base + ".png")

        try:
            im = PILImage.open(src_path).convert("RGB")
            w, h = im.size
            scale = self.thumb_width / float(w)
            im = im.resize((int(w * scale), int(h * scale)))
            im.save(out_path, "PNG")
            return out_path
        except Exception as e:
            print(f"⚠️ Thumbnail failed: {ErrorHandler.safe_err_str(e)}")
            return None

    # ====== Append row ======
    def append_row(self, order_idx, category, image_path, tv_str, av_str, zoom, iso_str, ec_val):
        """Append new log row with image and thumbnail."""
        timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
        row_data = [order_idx, timestamp, category, image_path, tv_str, av_str, str(zoom), iso_str, ec_val]
        self.ws.append(row_data)
        r = self.ws.max_row

        # Hyperlink
        try:
            target_url = "file:///" + image_path.replace("\\", "/")
            cell = self.ws.cell(row=r, column=4, value="open")
            cell.hyperlink = target_url
            cell.style = "Hyperlink"
        except Exception as e:
            print(f"⚠️ Failed hyperlink: {ErrorHandler.safe_err_str(e)}")

        # Thumbnail
        thumb = self._make_thumbnail(image_path)
        if thumb:
            try:
                img = XLImage(thumb)
                self.ws.add_image(img, f"D{r}")
                self.ws.row_dimensions[r].height = max(
                    self.ws.row_dimensions[r].height or 15, img.height * 0.75
                )
            except Exception as e:
                print(f"⚠️ Failed to insert thumbnail: {ErrorHandler.safe_err_str(e)}")

        try:
            self.wb.save(self.log_path)
            print(f"✅ Logged {os.path.basename(image_path)}")
        except PermissionError:
            alt = os.path.splitext(self.log_path)[0] + "_NEW.xlsx"
            self.wb.save(alt)
            print(f"⚠️ Excel was open — saved to {alt}")

    # ====== Reorder by FEATURE_ORDER ======
    def reorder_logfile(self):
        """Reorder Excel rows according to FEATURE_ORDER (category priority)."""
        try:
            wb = load_workbook(self.log_path)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=6, values_only=True))  # after metadata + header
            header = [c.value for c in ws[5]]

            order_map = {cat: i for i, cat in enumerate(self.feature_order)}
            rows.sort(key=lambda r: order_map.get(r[2], 999))  # column 3 = Feature

            ws.delete_rows(6, ws.max_row)
            for r in rows:
                ws.append(r)
                r_index = ws.max_row
                if len(r) >= 4 and isinstance(r[3], str):
                    cell = ws.cell(row=r_index, column=4, value="open")
                    link = "file:///" + r[3].replace("\\", "/")
                    try:
                        cell.hyperlink = link
                        cell.style = "Hyperlink"
                    except Exception:
                        cell.value = link

            wb.save(self.log_path)
            print("✅ Excel log reordered by FEATURE_ORDER.")
        except Exception as e:
            print(f"⚠️ Could not reorder log file: {ErrorHandler.safe_err_str(e)}")
