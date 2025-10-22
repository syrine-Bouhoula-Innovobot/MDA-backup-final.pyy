@ -0,0 +1,1003 @@
# Focus ONCE @ 140 with your requested sequence:
#   A) Half-press AF to wake (then RELEASE)
#   B) Robust ramp to zoom 140 (no half-press held)
#   C) Half-press AF AGAIN at 140, then capture NON-AF
#   D) Disable all AF (MF + kill continuous AF). From here on: NON-AF only.
#
# Common: Av f/8, ISO 100. Save to PC. Append to Excel log with thumbnails.

import os, sys, time, uuid, json

# ====== Load configuration ======
def load_config(config_path="camera_config.json"):
    """Load camera and category configuration from JSON file."""
    try:
        with open(config_path, "r") as f:
            cfg = json.load(f)
        print(f"Loaded configuration from {config_path}")
        return cfg
    except FileNotFoundError:
        print(f"Config file '{config_path}' not found. Using defaults.")
        return {}
    except Exception as e:
        print(f" Error reading config: {e}")
        return {}

CONFIG = load_config()

# ====== Project info ======
PROJECT_ID   = CONFIG.get("project", {}).get("PROJECT_ID", "MDA12345")
DEVICE_CODE  = CONFIG.get("project", {}).get("DEVICE_CODE", "DC123")

# ====== Camera parameters ======
camera_cfg = CONFIG.get("camera", {})
AV_LABEL     = camera_cfg.get("AV_LABEL", "f/8")
ISO_LABEL    = camera_cfg.get("ISO_LABEL", "100")
TV_REF_LABEL = camera_cfg.get("TV_REF_LABEL", "1/60")
DELAY_S      = camera_cfg.get("DELAY_S", 3.0)
POST_SHOT_WAIT     = camera_cfg.get("POST_SHOT_WAIT", 2.5)
THUMBNAIL_WIDTH_PX = camera_cfg.get("THUMBNAIL_WIDTH_PX", 180)
ZOOM_CFG     = camera_cfg.get("ZOOM_STEPS", {})

ZOOM_140_STR = ZOOM_CFG.get("ZOOM_140_STR", "140")
ZOOM_120_STR = ZOOM_CFG.get("ZOOM_120_STR", "120")
ZOOM_110_STR = ZOOM_CFG.get("ZOOM_110_STR", "110")
ZOOM_100_STR = ZOOM_CFG.get("ZOOM_100_STR", "100")
ZOOM_055_STR = ZOOM_CFG.get("ZOOM_055_STR", "55")

# ====== Category and order setup ======
CATS = CONFIG.get("categories", {})
def _cat(k, default=None): return CATS.get(k, default or k)

CAT_REF = _cat("CAT_REF", "reference focus sticker")

orders_cfg = CONFIG.get("orders", {})
FEATURE_ORDER = [_cat(k) for k in orders_cfg.get("FEATURE_ORDER", [])]
ORDER_Z140    = [_cat(k) for k in orders_cfg.get("ORDER_Z140", [])]
ORDER_Z120    = [_cat(k) for k in orders_cfg.get("ORDER_Z120", [])]
ORDER_Z055    = [_cat(k) for k in orders_cfg.get("ORDER_Z055", [])]
ORDER_Z110    = [_cat(k) for k in orders_cfg.get("ORDER_Z110", [])]

TV_MAP = {}
tv_map_cfg = CONFIG.get("tv_map", {})
for key, val in tv_map_cfg.items():
    TV_MAP[_cat(key)] = val

# ====== Environment and paths ======
HERE = os.path.dirname(os.path.abspath(__file__))
for p in [
    HERE,
    os.path.join(HERE, "EDSDK_64", "Dll"),
    os.path.join(HERE, "..", "EDSDK_64", "Dll"),
    os.path.join(HERE, "EDSDK", "Dll"),
    os.path.join(HERE, "..", "EDSDK", "Dll")
]:
    if hasattr(os, "add_dll_directory") and os.path.isdir(p):
        try:
            os.add_dll_directory(os.path.abspath(p))
        except Exception:
            pass

# ====== Imports ======
import edsdk
from edsdk import CameraCommand, PropID, SaveTo, ObjectEvent, \
                   FileCreateDisposition, Access, EdsObject

IS_WIN = (os.name == "nt")
if IS_WIN:
    import pythoncom

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

# ====== Tunables ======
START_SETTLE_SEC   = 1.2
WRITE_RETRIES      = 12
POST_ZOOM_SETTLE   = 0.7
AF_WAIT_SEC        = 1.1
DEFAULT_DELTA      = 3

# ====== Custom folder naming ======
timestamp    = time.strftime("%Y-%m-%d_%H-%M-%S")
session_dir  = os.path.join(HERE, f"{PROJECT_ID}_{timestamp}")
os.makedirs(session_dir, exist_ok=True)

CAPTURE_DIR = session_dir
log_filename = f"{PROJECT_ID}_{time.strftime('%Y-%m-%d')}_{DEVICE_CODE}.xlsx"
LOG_PATH     = os.path.join(session_dir, log_filename)

# ====== Globals ======
cam_glob = None
CURRENT_CATEGORY = CAT_REF
saved_paths = []
AF_DISABLED = False
OBJ_HANDLER_ENABLED = False


def set_category(cat: str):
    global CURRENT_CATEGORY
    CURRENT_CATEGORY = str(cat)

def pump(total=0.35, dt=0.05):
    t0 = time.time()
    while time.time() - t0 < total:
        try: edsdk.SendCommand(cam_glob, CameraCommand.ExtendShutDownTimer, 0)
        except Exception: pass
        if IS_WIN:
            try: pythoncom.PumpWaitingMessages()
            except Exception: pass
        time.sleep(dt)

def safe(cmd, *args):
    try: cmd(*args)
    except Exception: pass

def dc_zoom_id():
    try:    return getattr(edsdk, "kEdsPropID_DC_Zoom")
    except: return getattr(__import__("edsdk").PropID, "DC_Zoom")

def status_enum():
    try:
        from edsdk import StatusCommand
        return StatusCommand
    except Exception:
        return None

# ---- error helpers ----
def safe_err_str(e):
    try:
        return str(e)
    except Exception:
        try:
            if hasattr(e, "args") and e.args:
                return str(e.args[0])
        except Exception:
            pass
        try:
            return repr(e)
        except Exception:
            return "<unprintable error>"

def err_is_busy(e):
    s = safe_err_str(e).lower()
    return ("device_busy" in s) or ("eds_err_device_busy" in s) or ("busy" in s)

# ---------- LiveView keepalive (NO half-press) ----------
def keepalive_no_half(cam):
    safe(edsdk.SendCommand, cam, CameraCommand.ExtendShutDownTimer, 0)
    pump(0.15)
    try:
        cur = int(edsdk.GetPropertyData(cam, PropID.Evf_OutputDevice, 0))
        safe(edsdk.SetPropertyData, cam, PropID.Evf_OutputDevice, 0, cur)
    except Exception:
        pass
    pump(0.10)
    for pid in (PropID.Tv, PropID.Av, PropID.ISOSpeed):
        try: _ = edsdk.GetPropertyData(cam, pid, 0)
        except Exception: pass
    pump(0.10)

# ---- Zoom helpers ----
def get_desc(cam, pid):
    try:    return edsdk.GetPropertyDesc(cam, pid) or {}
    except Exception as e:
        print("GetPropertyDesc failed:", e); return {}

def derive_steps(desc):
    total, max_step = None, None
    if isinstance(desc, dict):
        pd = desc.get("propDesc")
        if isinstance(pd, (list, tuple)) and pd:
            if len(pd) == 1 and isinstance(pd[0], int):
                total = int(pd[0]); max_step = total - 1
            else:
                vals = [int(x) for x in pd if isinstance(x, int)]
                if vals:
                    max_step = max(vals); total = max_step - min(vals) + 1
        if total is None:
            for k in ("max","Max","maximum","Maximum"):
                if k in desc and isinstance(desc[k], int):
                    max_step = int(desc[k]); total = max_step + 1; break
    return total or 201, max_step or 200

def read_zoom(cam, pid):
    try: return int(edsdk.GetPropertyData(cam, pid, 0))
    except Exception: return None

def write_zoom(cam, pid, step, StatusCommand=None, retries=WRITE_RETRIES):
    step = int(step); last = ""
    for i in range(retries):
        try:
            edsdk.SetPropertyData(cam, pid, 0, step)
            pump(0.24 + 0.12*i)
            rb = read_zoom(cam, pid)
            if rb == step: return True
        except Exception as e:
            last = safe_err_str(e)
        pump(0.28 + 0.14*i)
    if last: print("Set DC_Zoom failed:", last)
    return False

def reset_liveview_for_zoom(cam):
    """Bootstrap LV so the FIRST ramp accepts writes."""
    try:
        safe(edsdk.SendCommand, cam, CameraCommand.PressShutterButton, 0)
        keepalive_no_half(cam)
        pump(0.20)
        try: safe(edsdk.SetPropertyData, cam, PropID.Evf_Mode, 0, 1)
        except Exception: pass
        pump(0.18)
        safe(edsdk.SetPropertyData, cam, PropID.Evf_OutputDevice, 0, 0)  # off
        pump(0.45)
        safe(edsdk.SetPropertyData, cam, PropID.Evf_OutputDevice, 0, 2)  # PC
        pump(1.0)
        for pid in (PropID.Tv, PropID.Av, PropID.ISOSpeed):
            try: _ = edsdk.GetPropertyData(cam, pid, 0)
            except Exception: pass
        pump(0.30)
    except Exception as e:
        print("reset_liveview_for_zoom: ignored error:", safe_err_str(e))

def ramp_zoom(cam, pid, target, delta=DEFAULT_DELTA, quiet_settle=0.0, first=False):
    """
    Robust zoom ramp that NEVER holds half-press.
    UI lock is held for the entire ramp to reduce BUSY flicker.
    If it fails at the beginning, we reset LV once and retry.
    """
    se = status_enum()

    def _do_ramp():
        # Make sure nothing is half-pressed.
        safe(edsdk.SendCommand, cam, CameraCommand.PressShutterButton, 0)
        keepalive_no_half(cam)
        pump(0.25)

        cur = read_zoom(cam, pid)
        if cur is None:
            print("ramp_zoom: cannot read zoom.")
            return False, "cannot read zoom"
        if target == cur:
            print(f"ramp_zoom: already at {cur}")
            pump(POST_ZOOM_SETTLE)
            return True, cur

        if se: safe(edsdk.SendStatusCommand, cam, se.UILock)
        try:
            sgn = 1 if target > cur else -1
            step = cur
            while step != target:
                step = step + sgn * min(delta, abs(target - step))
                ok = write_zoom(cam, pid, step, StatusCommand=None)
                print(f"  zoom write -> {step} : {'OK' if ok else 'FAIL'}")
                if not ok:
                    return False, f"write failed at step {step}"
                pump(0.08)
        finally:
            if se: safe(edsdk.SendStatusCommand, cam, se.UIUnLock)
        pump(POST_ZOOM_SETTLE)
        return True, step

    ok, msg = _do_ramp()
    if not ok:
        print("ramp_zoom: first attempt failed; resetting Live View and retrying…")
        reset_liveview_for_zoom(cam)
        pump(0.35)
        ok, msg = _do_ramp()

    if ok and quiet_settle > 0:
        print(f"Quiet settle {quiet_settle:.1f}s…")
        pump(quiet_settle)

    return ok, msg

# -------- Shutter helpers --------
SH_OFF        = 0x00000000
SH_HALF       = 0x00000001
SH_FULL_NONAF = 0x00010003  # full press without AF

def half_press_with_retry(cam, wait=AF_WAIT_SEC, retries=6):
    """Used for AF before we disable it. Busy-safe with backoff."""
    if AF_DISABLED:
        print("Half-press ignored (AF already disabled).")
        return False
    keepalive_no_half(cam)  # wake without half-press
    pump(0.25)
    for i in range(retries):
        try:
            edsdk.SendCommand(cam, CameraCommand.PressShutterButton, SH_HALF)
            pump(wait)
            return True
        except Exception as e:
            if err_is_busy(e):
                back = 0.35 + 0.25*i
                print(f"Half-press busy, retrying in {back:.2f}s…")
                pump(back)
                keepalive_no_half(cam)
                continue
            print("Half-press failed:", safe_err_str(e))
            return False
    print("Half-press failed: device busy (exhausted retries).")
    return False

def release_shutter(cam):
    safe(edsdk.SendCommand, cam, CameraCommand.PressShutterButton, SH_OFF)

def capture_full_nonaf(cam, hold=0.5, retries=6):
    for i in range(retries):
        try:
            edsdk.SendCommand(cam, CameraCommand.PressShutterButton, SH_FULL_NONAF)
            pump(hold)
            release_shutter(cam)
            return True
        except Exception as e:
            if err_is_busy(e):
                pump(0.25 + 0.25*i)
                safe(edsdk.SendCommand, cam, CameraCommand.ExtendShutDownTimer, 0)
                continue
            print("NON-AF capture failed:", safe_err_str(e))
            release_shutter(cam)
            return False
    print("NON-AF capture failed: device busy.")
    release_shutter(cam)
    return False

def focus_once_and_lock(cam):
    """Perform one half-press AF, then disable all autofocus (PowerShot-safe)."""
    print("\n Running focus-lock sequence before shooting...")
    ok = half_press_with_retry(cam, AF_WAIT_SEC)
    if ok:
        print("Focus acquired.")
    else:
        print(" Focus attempt failed or was skipped.")
    release_shutter(cam)
    pump(0.4)

    print(" Disabling autofocus for the rest of the session...")
    try:
        for name in ("ContinuousAFMode", "MovieServoAF", "LensDriveWhenAFImpossible", "AFAssist"):
            pid = getattr(PropID, name, None)
            if pid:
                try:
                    edsdk.SetPropertyData(cam, pid, 0, 0)
                    print(f"{name}=Off")
                except Exception:
                    pass
        try:
            cur = int(edsdk.GetPropertyData(cam, PropID.Evf_OutputDevice, 0))
            edsdk.SetPropertyData(cam, PropID.Evf_OutputDevice, 0, cur)
        except Exception:
            pass
        print(" Autofocus locked. Camera will stay fixed.")
    except Exception as e:
        print("AF disable error:", e)

def release_all_locks(cam):
    release_shutter(cam)

# -------- Save-to-PC callback + fallback ------------------------------------
def ensure_dir(path): os.makedirs(path, exist_ok=True)
def category_dir():
    return CAPTURE_DIR

def _save_item(handle: EdsObject, out_dir: str):
    info = edsdk.GetDirectoryItemInfo(handle)
    try:
        cat_prefix = f"{PROJECT_ID}_{CURRENT_CATEGORY.replace(' ', '_')}"
        base_name  = f"{cat_prefix}_{time.strftime('%H%M%S')}"
        out_path   = os.path.join(out_dir, base_name + ".JPG")

        stream = edsdk.CreateFileStream(out_path, FileCreateDisposition.CreateAlways, Access.ReadWrite)
        edsdk.Download(handle, info["size"], stream)
        edsdk.DownloadComplete(handle)
        saved_paths.append(out_path)
        print("Saved:", out_path)
    except Exception as e:
        print("Save failed:", e)

def obj_callback(event: ObjectEvent, handle: EdsObject) -> int:
    try:
        if event in (ObjectEvent.DirItemRequestTransfer, ObjectEvent.DirItemCreated):
            _save_item(handle, category_dir())
    except Exception as e:
        print("Save (callback) failed:", safe_err_str(e))
    return 0

def enable_object_handler_after_ref(cam):
    global OBJ_HANDLER_ENABLED
    if OBJ_HANDLER_ENABLED: return
    try:
        edsdk.SetObjectEventHandler(cam, ObjectEvent.All, obj_callback)
        OBJ_HANDLER_ENABLED = True
        print("ObjectEvent handler enabled AFTER reference shot.")
    except Exception as e:
        print("ObjectEvent handler enable failed:", safe_err_str(e))

def list_jpgs(path):
    if not os.path.isdir(path): return []
    return [os.path.join(path, f) for f in os.listdir(path)
            if f.lower().endswith((".jpg", ".jpeg"))]

# ---------------- Tv / Av / ISO helpers ----------------
TV_LABELS = {24:"30\"",27:"25\"",29:"20\"",32:"15\"",35:"13\"",37:"10\"",40:"8\"",43:"6\"",45:"5\"",
             48:"4\"",51:"3\"2",53:"2\"5",56:"2\"",59:"1\"6",61:"1\"3",64:"1\"",67:"0\"8",69:"0\"6",
             72:"0\"5",75:"0\"4",77:"0\"3",80:"1/5",83:"1/6",85:"1/8",88:"1/10",
             91:"1/20",93:"1/25",96:"1/30",99:"1/40",101:"1/50",104:"1/60",107:"1/80",
             109:"1/100",112:"1/125",115:"1/160",117:"1/200",120:"1/250",123:"1/320",
             125:"1/400",128:"1/500",131:"1/640",133:"1/800",136:"1/1000",139:"1/1250",
             141:"1/1600",144:"1/2000"}
def tv_label(code): return TV_LABELS.get(int(code), f"code={int(code)}")
def tv_desc(cam):
    try: return [int(v) for v in (edsdk.GetPropertyDesc(cam, PropID.Tv) or {}).get("propDesc") or []]
    except Exception: return []
def parse_label_to_index(vals, want_label):
    want = want_label.strip()
    for i, c in enumerate(vals):
        if tv_label(c) == want or str(c) == want: return i
    raise ValueError("Tv label not in allowed list")
def read_tv(cam):
    try: return int(edsdk.GetPropertyData(cam, PropID.Tv, 0))
    except Exception: return None
def set_tv_by_index(cam, idx, vals):
    code = int(vals[idx]); edsdk.SetPropertyData(cam, PropID.Tv, 0, code); pump(0.6)
    rb = read_tv(cam)
    print(f"Tv: set code={code} ({tv_label(code)}) -> readback {rb} ({tv_label(rb)})")
    return code, rb
def set_tv_preferred(cam, want_label):
    vals = tv_desc(cam)
    if not vals:
        print("No Tv list; leaving Tv as-is.")
        return None, read_tv(cam)
    lookup = {"1/25":93,"1/20":91,"1/100":109,"1/30":96,"1/50":101,"1/80":107,"1/40":99,"1/125":112,"1/320":123}
    try:
        idx = parse_label_to_index(vals, want_label); return set_tv_by_index(cam, idx, vals)
    except Exception:
        target_code = lookup.get(want_label,93)
        idx = min(range(len(vals)), key=lambda i: abs(int(vals[i])-target_code))
        return set_tv_by_index(cam, idx, vals)

AV_LABELS = {40:"f/3.5",43:"f/4.0",45:"f/4.5",48:"f/5.0",51:"f/5.6",53:"f/7.1",56:"f/8",59:"f/9.0",61:"f/10",64:"f/11",67:"f/13"}
def av_label(code): return AV_LABELS.get(int(code), f"code={int(code)}")
def av_desc(cam):
    try: return [int(v) for v in (edsdk.GetPropertyDesc(cam, PropID.Av) or {}).get("propDesc") or []]
    except Exception: return []
def choose_av_code(vals, want_label):
    want = (want_label or "").replace(".0","")
    for c in vals:
        if av_label(c).replace(".0","") == want: return int(c)
    return int(min(vals, key=lambda c: abs(int(c)-56))) if vals else None
def set_av(cam, code):
    edsdk.SetPropertyData(cam, PropID.Av, 0, int(code)); pump(0.5)
    rb = int(edsdk.GetPropertyData(cam, PropID.Av, 0))
    print(f"Av: set code={code} ({av_label(code)}) -> readback {rb} ({av_label(rb)})"); return rb

def _iso_desc(cam):
    try:
        d = edsdk.GetPropertyDesc(cam, PropID.ISOSpeed) or {}
        return [int(v) for v in d.get("propDesc") or []]
    except Exception: return []
_STD_ISO_SERIES = [100,125,160,200,250,320,400,500,640,800,1000,1250,1600,2000,2500,3200]
def _iso_series_for_desc(desc):
    nz = [c for c in desc if c != 0]
    return dict(zip(nz, _STD_ISO_SERIES[:len(nz)]))
def iso_label_from_code(cam, code):
    desc = _iso_desc(cam)
    if not desc: return "Auto" if code == 0 else f"code={int(code)}"
    if code == 0: return "Auto"
    return str(_iso_series_for_desc(desc).get(int(code), f"code={int(code)}"))
def set_iso_preferred(cam, want_label: str):
    desc = _iso_desc(cam)
    if not desc:
        print("ISO list not exposed; skipping ISO set.")
        try: return int(edsdk.GetPropertyData(cam, PropID.ISOSpeed, 0))
        except Exception: return None
    map_by_order = _iso_series_for_desc(desc); inv = {v:k for k,v in map_by_order.items()}
    s = str(want_label).strip().lower().replace("iso","")
    if s in ("auto","0"): target_code = 0 if 0 in desc else None
    else:
        try: want_num = int(s)
        except Exception: want_num = None
        if want_num in inv: target_code = inv[want_num]
        else:
            if want_num is None or not inv: target_code = None
            else:
                nearest = min(map_by_order.values(), key=lambda x: abs(x - want_num))
                target_code = inv[nearest]
    if target_code is None:
        print("ISO: could not resolve requested value; leaving as-is.")
        try: return int(edsdk.GetPropertyData(cam, PropID.ISOSpeed, 0))
        except Exception: return None
    edsdk.SetPropertyData(cam, PropID.ISOSpeed, 0, int(target_code)); pump(0.5)
    rb = int(edsdk.GetPropertyData(cam, PropID.ISOSpeed, 0))
    print(f"ISO: set code={target_code} ({iso_label_from_code(cam, target_code)}) -> "
          f"readback {rb} ({iso_label_from_code(cam, rb)})")
    return rb
def read_iso_label(cam):
    try:
        code = int(edsdk.GetPropertyData(cam, PropID.ISOSpeed, 0))
        return iso_label_from_code(cam, code)
    except Exception: return ""

# -------- EC as numeric EV --------
def _ec_desc_list(cam):
    try: return [int(v) for v in (edsdk.GetPropertyDesc(cam, PropID.ExposureCompensation) or {}).get("propDesc") or []]
    except Exception: return []
def read_ec(cam):
    try: cur_code = int(edsdk.GetPropertyData(cam, PropID.ExposureCompensation, 0))
    except Exception: return ""
    vals = _ec_desc_list(cam)
    if not vals or cur_code not in vals: return 0.0 if cur_code == 0 else ""
    z = vals.index(0); i = vals.index(cur_code); steps = i - z
    whole, rem = divmod(abs(steps), 3); frac = {0:0.0,1:0.3,2:0.7}[rem]
    v = (whole + frac) * (1 if steps >= 0 else -1)
    return float(f"{max(-3.0, min(3.0, v)):.1f}")

# -------- Excel logging --------
HEADERS = ["timestamp","picture category","picture","hyperlink","tv","av","zoom_step","iso","ec"]
_RESOLVED_LOG_PATH = None

def _detect_or_upgrade_log_path():
    if os.path.exists(LOG_PATH):
        try:
            wb = load_workbook(LOG_PATH)
            ws = wb.active
            first = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            if first == HEADERS:
                print("Using existing log file:", LOG_PATH)
                return LOG_PATH
        except Exception:
            pass
    wb = Workbook()
    ws = wb.active
    ws.append(HEADERS)
    wb.save(LOG_PATH)
    print("Created new log file:", LOG_PATH)
    return LOG_PATH

def _get_log_path():
    global _RESOLVED_LOG_PATH
    if _RESOLVED_LOG_PATH is None:
        _RESOLVED_LOG_PATH = _detect_or_upgrade_log_path()
    return _RESOLVED_LOG_PATH

def tv_text_for_log(tv_str):
    return f"{tv_str} s" if tv_str and "/" in tv_str and '"' not in tv_str else tv_str

def make_png_thumbnail(src_path, out_dir):
    if not (PILImage and os.path.exists(src_path)): return None
    os.makedirs(out_dir, exist_ok=True)
    base = os.path.splitext(os.path.basename(src_path))[0]
    out_path = os.path.join(out_dir, base + ".png")
    try:
        im = PILImage.open(src_path)
        try: im.seek(0)
        except Exception: pass
        im = im.convert("RGB"); w,h = im.size; scale = THUMBNAIL_WIDTH_PX/float(w)
        im = im.resize((int(w*scale), int(h*scale))); im.save(out_path, "PNG")
        return out_path
    except Exception as e:
        print("Thumbnail conversion failed:", safe_err_str(e)); return None

def append_xlsx_row(row, image_path=None):
    xlsx_path = _get_log_path()
    try:
        wb = load_workbook(xlsx_path); ws = wb.active
    except Exception:
        wb = Workbook(); ws = wb.active; ws.append(HEADERS)
    try: ws.column_dimensions["C"].width = 28
    except Exception: pass

    ws.append(row); r = ws.max_row
    ws.cell(row=r, column=3).value = ""  # thumbnail column

    target_url = "file:///" + row[3].replace('\\','/')
    cell = ws.cell(row=r, column=4, value="open")
    try: cell.hyperlink = target_url; cell.style = "Hyperlink"
    except Exception: cell.value = target_url

    if image_path:
        thumb_png = make_png_thumbnail(image_path, os.path.join(CAPTURE_DIR, "_thumbs"))
        if thumb_png:
            try:
                img = XLImage(thumb_png); ws.add_image(img, f"C{r}")
                ws.row_dimensions[r].height = max(ws.row_dimensions[r].height or 15,
                                                  img.height * 0.75)
            except Exception as e:
                print("Thumbnail add failed:", safe_err_str(e))

    try: wb.save(xlsx_path)
    except PermissionError:
        alt = os.path.splitext(xlsx_path)[0] + "_NEW.xlsx"
        print("XLSX in use; writing to", alt); wb.save(alt)
        global _RESOLVED_LOG_PATH
        _RESOLVED_LOG_PATH = alt
    return xlsx_path

def write_excel_metadata(ws, header_cfg):
    """Write project metadata rows at the top of the Excel sheet."""
    ws["A1"] = "Part number"
    ws["B1"] = header_cfg.get("Part number", "")
    ws["A2"] = "Part Description"
    ws["B2"] = header_cfg.get("Part Description", "")
    ws["A3"] = "Serial number"
    ws["B3"] = header_cfg.get("Serial number", "")
    ws["A4"] = "Program type"
    ws["B4"] = header_cfg.get("Program type", "")

def _detect_or_upgrade_log_path():
    """Create new Excel log file (or reuse if exists) with header metadata."""
    excel_header = CONFIG.get("excel_header", {})
    if os.path.exists(LOG_PATH):
        try:
            wb = load_workbook(LOG_PATH)
            ws = wb.active
            first = [c.value for c in next(ws.iter_rows(min_row=5, max_row=5))]
            if first and "timestamp" in str(first):
                print("Using existing log file:", LOG_PATH)
                return LOG_PATH
        except Exception:
            pass

    wb = Workbook()
    ws = wb.active
    # Insert metadata lines
    write_excel_metadata(ws, excel_header)
    # Column headers (start from row 5)
    ws.append(["Order", "timestamp", "Feature", "Image", "tv", "av", "zoom", "iso", "ec"])
    wb.save(LOG_PATH)
    print("Created new log file with metadata:", LOG_PATH)
    return LOG_PATH

# -------------------- Shot pipeline --------------------
def do_shot(cam, *, category, tv_label_use, is_reference, zoom_pid, zoom_target, av_label_use, iso_label_use):
    set_category(category)

    folder_before = set(list_jpgs(category_dir()))

    # Tv first (quiet)
    _res = set_tv_preferred(cam, tv_label_use)
    tv_code = _res[1] if isinstance(_res, tuple) else read_tv(cam)
    tv_str = tv_label(tv_code) if tv_code is not None else ""

    # Zoom (if requested)
    if zoom_target is not None:
        desc = get_desc(cam, zoom_pid); total, max_step = derive_steps(desc)
        cur = read_zoom(cam, zoom_pid)
        print(f"DC_Zoom: total={total}, range 0..{max_step}, current={cur}, target={zoom_target}")
        ok, msg = ramp_zoom(cam, zoom_pid, zoom_target,
                            delta=DEFAULT_DELTA, quiet_settle=0.0, first=False)
        if not ok: print("Zoom ramp FAILED:", msg)
        rbz = read_zoom(cam, zoom_pid)
        print("Zoom readback:", rbz)
    else:
        rbz = read_zoom(cam, zoom_pid)
        print("Zoom unchanged at:", rbz)

    # Av / ISO
    av_vals = av_desc(cam)
    if av_vals:
        av_code = choose_av_code(av_vals, av_label_use)
        if av_code is not None: set_av(cam, av_code)
    try: av_code_now = int(edsdk.GetPropertyData(cam, PropID.Av, 0))
    except Exception: av_code_now = None
    av_str = av_label(av_code_now) if av_code_now is not None else ""

    iso_code_now = set_iso_preferred(cam, iso_label_use)
    iso_str = iso_label_from_code(cam, iso_code_now) if iso_code_now is not None else read_iso_label(cam)

    # Shoot
    start_idx = len(saved_paths)
    pump(0.25)
    if is_reference:
        print("Reference capture (NON-AF)…")
        capture_full_nonaf(cam)
    else:
        print("Capture (NON-AF)…")
        capture_full_nonaf(cam)

    # Wait + gather new files
    pump(POST_SHOT_WAIT)
    new_files = saved_paths[start_idx:]

    # Fallback scan
    if not new_files:
        folder_after = set(list_jpgs(category_dir()))
        diff = sorted(folder_after - folder_before, key=os.path.getmtime)
        if diff:
            print(f"Fallback picked up {len(diff)} file(s).")
            new_files = diff
            saved_paths.extend([p for p in diff if p not in saved_paths])

    ec_val = read_ec(cam)

    if new_files:
        for path in new_files:
            ts = time.strftime("%Y-%m-%d %H:%M:%S")
            row = [ts, category, os.path.basename(path), path,
                   tv_text_for_log(tv_str), av_str, str(rbz if rbz is not None else ""),
                   iso_str, ec_val]
            append_xlsx_row(row, image_path=path)
        print(f"Logged {len(new_files)} shot(s) for '{category}'.")
    else:
        print("No files detected to log for this shot.")

# -------------------- EVF --------------------
def enable_evf(cam):
    safe(edsdk.SetPropertyData, cam, PropID.Evf_Mode, 0, 1); pump(0.18)
    safe(edsdk.SetPropertyData, cam, PropID.Evf_OutputDevice, 0, 2); pump(0.22)
    print("EVF mode -> ON\nEVF output -> PC")

def reorder_logfile():
    from openpyxl import load_workbook
    wb = load_workbook(LOG_PATH)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    header = [c.value for c in ws[1]]

    # Sort rows based on desired FEATURE_ORDER
    order_map = {cat: i for i, cat in enumerate(FEATURE_ORDER)}
    rows.sort(key=lambda r: order_map.get(r[1], 999))

    # Rewrite sheet and restore hyperlinks
    ws.delete_rows(2, ws.max_row)
    for r in rows:
        ws.append(r)
        r_index = ws.max_row
        if len(r) >= 4 and isinstance(r[3], str):
            cell = ws.cell(row=r_index, column=4, value="open")
            link = "file:///" + r[3].replace("\\", "/")
            try:
                cell.hyperlink = link
                cell.style = "Hyperlink"
            except Exception:
                cell.value = link

    wb.save(LOG_PATH)
    print("Excel log reordered by FEATURE_ORDER.")

def disable_all_af_safe(cam):
    """Disable all AF systems (PowerShot-safe)."""
    print("\n Disabling autofocus (PowerShot-safe)…")
    try:
        from edsdk import StatusCommand
        se = StatusCommand
    except Exception:
        se = None

    def set_prop(pid_name, value, label=""):
        pid = getattr(PropID, pid_name, None)
        if pid is None:
            return
        if se:
            safe(edsdk.SendStatusCommand, cam, se.UILock)
        try:
            edsdk.SetPropertyData(cam, pid, 0, value)
            if label:
                print(f"{label} -> {value}")
        except Exception as e:
            if label:
                print(f"{label} set failed:", safe_err_str(e))
        finally:
            if se:
                safe(edsdk.SendStatusCommand, cam, se.UIUnLock)
        pump(0.15)

    props = [
        ("ContinuousAFMode", 0, "ContinuousAFMode=Off"),
        ("MovieServoAF", 0, "MovieServoAF=Off"),
        ("LensDriveWhenAFImpossible", 0, "LensDriveWhenAFImpossible=Off"),
        ("AFAssist", 0, "AFAssist=Off"),
    ]
    for n, v, l in props:
        set_prop(n, v, l)

    # Refresh EVF to apply “pseudo-MF” freeze
    try:
        cur = int(edsdk.GetPropertyData(cam, PropID.Evf_OutputDevice, 0))
        edsdk.SetPropertyData(cam, PropID.Evf_OutputDevice, 0, cur)
    except Exception:
        pass
    print(" AF systems off. Focus should remain fixed.\n")

def reset_af_state(cam):
    """Re-enable all autofocus systems and clear any half-press lock before the main sequence."""
    print("\n Resetting AF / Focus system (unlock half-press)…")

    # Enable AF-related properties
    for pid_name in ("ContinuousAFMode", "MovieServoAF", "LensDriveWhenAFImpossible", "AFAssist"):
        pid = getattr(PropID, pid_name, None)
        if pid is None:
            continue
        try:
            edsdk.SetPropertyData(cam, pid, 0, 1)
            print(f"{pid_name} → 1 (enabled)")
        except Exception as e:
            print(f"Could not enable {pid_name}:", safe_err_str(e))

    # Set AF mode to One-Shot (0) if supported
    try:
        edsdk.SetPropertyData(cam, PropID.AFMode, 0, 0)
        print("AFMode → 0 (One-Shot AF)")
    except Exception as e:
        print("Could not set AFMode:", safe_err_str(e))

    # Refresh EVF and run quick AF pulse
    try:
        edsdk.SetPropertyData(cam, PropID.Evf_Mode, 0, 1)
        edsdk.SetPropertyData(cam, PropID.Evf_OutputDevice, 0, 2)
        edsdk.SendCommand(cam, CameraCommand.DoEvfAf, 1)
        pump(0.5)
        edsdk.SendCommand(cam, CameraCommand.DoEvfAf, 0)
        print(" AF pulse sent to re-initialize lens drive.")
    except Exception:
        pass

    print(" AF state unlocked — half-press should now work normally.\n")

# -------------------- Main --------------------
def main():
    print(" Initializing Canon SDK…")
    edsdk.InitializeSDK()
    pid_zoom = dc_zoom_id()

    try:
        # ---------- CAMERA CONNECTION ----------
        cl = edsdk.GetCameraList()
        if edsdk.GetChildCount(cl) == 0:
            print(" No camera detected.")
            return 1
        cam = edsdk.GetChildAtIndex(cl, 0)
        global cam_glob
        cam_glob = cam

        print(" Camera connected. Opening session…")
        edsdk.OpenSession(cam)
        pump(1.0)

        # --- UNLOCK AF STATE ---
        reset_af_state(cam)
        pump(0.6)

        # ---------- STORAGE SETTINGS ----------
        edsdk.SetPropertyData(cam, PropID.SaveTo, 0, SaveTo.Host)
        edsdk.SetCapacity(cam, {
            "reset": True,
            "bytesPerSector": 512,
            "numberOfFreeClusters": 2_147_483_647
        })
        print(" SaveTo=Host + capacity set.")

        # ---------- ENABLE EVF ----------
        enable_evf(cam)

        # === A) Half-press AF to wake ===
        print("\n Step A: Half-press AF to wake…")
        if half_press_with_retry(cam, AF_WAIT_SEC):
            print(" Camera awake and focused (initial).")
        else:
            print(" Initial half-press failed.")
        release_shutter(cam)
        pump(0.4)

        # === B) Robust ramp to zoom 140 ===
        desc = get_desc(cam, pid_zoom)
        _, max_step = derive_steps(desc)
        zoom_140 = min(int(ZOOM_140_STR), max_step)
        print(f"\n Step B: Ramping to zoom {zoom_140} (no half-press held)…")
        ok, msg = ramp_zoom(cam, pid_zoom, zoom_140, quiet_settle=0.8, first=True)
        if ok:
            print(" Zoom ramp complete.")
        else:
            print(" Zoom ramp failed:", msg)
        print(" Waiting 1.2 s for lens settle…")
        time.sleep(1.2)

        # === C) Half-press AF AGAIN at 140, then lock ===
        print("\n Step C: Half-press AF again at zoom 140…")
        if half_press_with_retry(cam, wait=1.8):
            print(" Focus acquired at zoom 140.")
        else:
            print(" Focus attempt failed at zoom 140.")
        release_shutter(cam)
        pump(0.3)

        # === D) Disable all AF (manual-focus lock) ===
        print("\n Step D: Disabling all autofocus modes (locking focus)…")
        disable_all_af_safe(cam)
        try:
            edsdk.SetPropertyData(cam, PropID.AFMode, 0, 3)  # Manual Focus (if supported)
            print("AFMode → 3 (Manual Focus).")
        except Exception:
            pass
        try:
            cur_ev = int(edsdk.GetPropertyData(cam, PropID.Evf_OutputDevice, 0))
            edsdk.SetPropertyData(cam, PropID.Evf_OutputDevice, 0, cur_ev)
            print(" EVF refreshed to freeze focus state.")
        except Exception:
            pass
        print(" Focus locked — all subsequent shots will be NON-AF.\n")

        # ---------- ENABLE FILE TRANSFER CALLBACK ----------
        enable_object_handler_after_ref(cam)

        # ---------- MANUAL CAPTURE SEQUENCE ----------
        print("\n--- Begin manual capture sequence ---")
        print("Press ENTER when ready for each picture.\n")

        # ---- ZOOM 140 GROUP ----
        print("\n--- ZOOM 140 group ---")
        for cat in ORDER_Z140:
            input(f"\nReady for '{cat}' — press ENTER to capture…")
            do_shot(
                cam,
                category=cat,
                tv_label_use=TV_MAP.get(cat, TV_REF_LABEL),
                is_reference=(cat == CAT_REF),
                zoom_pid=pid_zoom,
                zoom_target=None,
                av_label_use=AV_LABEL,
                iso_label_use=ISO_LABEL
            )

        # ---- ZOOM 120 GROUP ----
        print("\n--- ZOOM 120 group ---")
        for idx, cat in enumerate(ORDER_Z120):
            input(f"\nReady for '{cat}' — press ENTER to capture…")
            do_shot(
                cam,
                category=cat,
                tv_label_use=TV_MAP.get(cat, TV_REF_LABEL),
                is_reference=False,
                zoom_pid=pid_zoom,
                zoom_target=120 if idx == 0 else None,
                av_label_use=AV_LABEL,
                iso_label_use=ISO_LABEL
            )

        # ---- ZOOM 55 GROUP ----
        print("\n--- ZOOM 55 group ---")
        for idx, cat in enumerate(ORDER_Z055):
            input(f"\nReady for '{cat}' — press ENTER to capture…")
            do_shot(
                cam,
                category=cat,
                tv_label_use=TV_MAP.get(cat, TV_REF_LABEL),
                is_reference=False,
                zoom_pid=pid_zoom,
                zoom_target=55 if idx == 0 else None,
                av_label_use=AV_LABEL,
                iso_label_use=ISO_LABEL
            )

        # ---- ZOOM 110 GROUP (LAST) ----
        print("\n--- ZOOM 110 group (LAST) ---")
        for cat in ORDER_Z110:
            input(f"\nReady for '{cat}' — press ENTER to capture…")
            do_shot(
                cam,
                category=cat,
                tv_label_use=TV_MAP.get(cat, TV_REF_LABEL),
                is_reference=False,
                zoom_pid=pid_zoom,
                zoom_target=110,
                av_label_use=AV_LABEL,
                iso_label_use=ISO_LABEL
            )

        # ---------- FINALIZE ----------
        reorder_logfile()
        print("\n All captures complete.")
        print(" Files saved under:", CAPTURE_DIR)
        print(" Log file used:", _get_log_path())

    finally:
        # ---------- SAFE EXIT ----------
        release_all_locks(cam_glob)
        print(" Closing session…")
        safe(edsdk.CloseSession, cam_glob)
        print(" Terminating SDK…")
        edsdk.TerminateSDK()
        pump(0.6)

if __name__ == "__main__":
    sys.exit(main() or 0)
